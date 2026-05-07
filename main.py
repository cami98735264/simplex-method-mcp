"""Simplex Method solver — JSON in, Excel out.

Reproduces the step-by-step teaching style from the reference PDFs
(Concept / Problem 1 / Problem 2): every tableau is preserved,
the pivot column / row / element are highlighted, the Cj-Zj row keeps
Big-M expressions symbolic, and a final SOLUCIÓN ÓPTIMA block lists
each variable's value.
"""
from __future__ import annotations

import json
import sys
from dataclasses import dataclass, field
from fractions import Fraction
from pathlib import Path
from typing import Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# Section 1 — BigM: exact (constant + m_coef · M) arithmetic
# =============================================================================
class BigM:
    """A value of the form c + m·M, where M is a symbolic large positive."""

    __slots__ = ("c", "m")

    def __init__(self, c: Union[int, Fraction], m: Union[int, Fraction] = 0):
        self.c = Fraction(c)
        self.m = Fraction(m)

    @staticmethod
    def coerce(other) -> "BigM":
        if isinstance(other, BigM):
            return other
        if isinstance(other, (int, Fraction)):
            return BigM(other, 0)
        raise TypeError(f"cannot coerce {type(other).__name__} to BigM")

    def __add__(self, other):
        o = BigM.coerce(other)
        return BigM(self.c + o.c, self.m + o.m)

    def __radd__(self, other):
        return BigM.coerce(other) + self

    def __sub__(self, other):
        o = BigM.coerce(other)
        return BigM(self.c - o.c, self.m - o.m)

    def __rsub__(self, other):
        return BigM.coerce(other) - self

    def __mul__(self, other):
        o = BigM.coerce(other)
        # (c1 + m1 M)(c2 + m2 M) — we only ever multiply BigM by a pure scalar
        # in this solver (Cj_basic · column_entry where column entries are pure
        # fractions), so reject the M·M case.
        if self.m != 0 and o.m != 0:
            raise ArithmeticError("M · M term is not supported")
        return BigM(self.c * o.c, self.c * o.m + self.m * o.c)

    def __rmul__(self, other):
        return self * other

    def __truediv__(self, other):
        # division only by a pure scalar
        if isinstance(other, BigM):
            if other.m != 0:
                raise ArithmeticError("division by a non-scalar BigM is not supported")
            other = other.c
        d = Fraction(other)
        return BigM(self.c / d, self.m / d)

    def __neg__(self):
        return BigM(-self.c, -self.m)

    def __eq__(self, other):
        try:
            o = BigM.coerce(other)
        except TypeError:
            return NotImplemented
        return self.c == o.c and self.m == o.m

    def __hash__(self):
        return hash((self.c, self.m))

    def __lt__(self, other):
        o = BigM.coerce(other)
        # M is treated as +∞: order by m first, then by c.
        if self.m != o.m:
            return self.m < o.m
        return self.c < o.c

    def __le__(self, other):
        return self == other or self < other

    def __gt__(self, other):
        return not self <= other

    def __ge__(self, other):
        return not self < other

    def is_zero(self) -> bool:
        return self.c == 0 and self.m == 0

    def __repr__(self):
        return f"BigM({self.c}, {self.m})"


# =============================================================================
# Section 2 — Formatting helpers
# =============================================================================
def fmt_fraction(f: Union[int, Fraction]) -> str:
    """Render a Fraction as integer, simple fraction, or mixed fraction."""
    f = Fraction(f)
    if f == 0:
        return "0"
    if f.denominator == 1:
        return str(f.numerator)
    sign = "-" if f < 0 else ""
    n, d = abs(f.numerator), f.denominator
    whole, rem = divmod(n, d)
    if whole == 0:
        return f"{sign}{rem}/{d}"
    return f"{sign}{whole} {rem}/{d}"


def _format_linear_expr(terms: list[tuple], *, keep_zero: bool) -> str:
    """Render a sum like '3·X1 - X2 + M·A2' from (coeff, label) pairs.

    Drops zero terms when `keep_zero` is False; turns '+ -1·v' into '- v',
    '+ 1·v' into '+ v', and the leading '1·v' into 'v'.
    """
    pieces: list[str] = []
    for coeff, label in terms:
        is_bigm = isinstance(coeff, BigM)
        is_zero = (coeff.is_zero() if is_bigm else coeff == 0)
        if is_zero and not keep_zero:
            continue

        # Determine sign for the connector and the magnitude string.
        if is_bigm:
            negate = (coeff.m < 0) or (coeff.m == 0 and coeff.c < 0)
            magnitude = (-coeff) if negate else coeff
            mag_str = fmt_bigm(magnitude)
            is_one = (magnitude.c == 1 and magnitude.m == 0)
        else:
            negate = coeff < 0
            magnitude = -coeff if negate else coeff
            mag_str = fmt_fraction(magnitude)
            is_one = (magnitude == 1)

        connector = "-" if negate else "+"
        body = label if is_one else f"{mag_str}·{label}"

        if not pieces:
            pieces.append(f"-{body}" if negate else body)
        else:
            pieces.append(f" {connector} {body}")

    return "".join(pieces) if pieces else "0"


def fmt_bigm(b: Union[BigM, Fraction, int]) -> str:
    """Render a BigM (or scalar) as 'a + bM' / 'a - bM' / 'M' / 'a' / etc."""
    if not isinstance(b, BigM):
        return fmt_fraction(b)
    c, m = b.c, b.m
    if m == 0:
        return fmt_fraction(c)
    # Build the M term
    if m == 1:
        m_term = "M"
    elif m == -1:
        m_term = "-M"
    else:
        m_term = f"{fmt_fraction(m)}M"
    if c == 0:
        return m_term
    # Combine: prefer "M - 1/3" over "-1/3 + M" for readability when m > 0
    if m > 0:
        m_part = "M" if m == 1 else f"{fmt_fraction(m)}M"
        if c > 0:
            return f"{fmt_fraction(c)} + {m_part}"
        return f"{m_part} - {fmt_fraction(-c)}"
    # m < 0
    abs_m = -m
    m_part = "M" if abs_m == 1 else f"{fmt_fraction(abs_m)}M"
    # Same template works whether c is positive or negative since
    # fmt_fraction(c) already carries the sign.
    return f"{fmt_fraction(c)} - {m_part}"


# =============================================================================
# Section 3 — JSON schema and Problem loader
# =============================================================================
@dataclass
class Constraint:
    name: str
    coefficients: list[Fraction]
    sign: str        # "<=", "=", ">="
    rhs: Fraction


@dataclass
class Problem:
    name: str
    objective: str   # "max" or "min"
    variables: list[str]
    objective_coefficients: list[Fraction]
    constraints: list[Constraint]


def _to_fraction(value, where: str) -> Fraction:
    if isinstance(value, bool):
        raise ValueError(f"{where}: boolean is not a valid number")
    if isinstance(value, (int, float)):
        return Fraction(value).limit_denominator(10**9)
    if isinstance(value, str):
        try:
            return Fraction(value)
        except (ValueError, ZeroDivisionError) as e:
            raise ValueError(f"{where}: cannot parse '{value}' as a number") from e
    raise ValueError(f"{where}: expected a number, got {type(value).__name__}")


def load_problem(path: Path) -> Problem:
    """Parse + validate a JSON LP problem file."""
    raw = json.loads(path.read_text(encoding="utf-8"))

    for key in ("objective", "variables", "objective_coefficients", "constraints"):
        if key not in raw:
            raise ValueError(f"missing required key: '{key}'")

    name = str(raw.get("name", path.stem))
    objective = str(raw["objective"]).lower()
    if objective not in ("max", "min"):
        raise ValueError(f"objective must be 'max' or 'min', got '{raw['objective']}'")

    variables = list(raw["variables"])
    if not variables:
        raise ValueError("variables list is empty")

    obj_coeffs_raw = raw["objective_coefficients"]
    if len(obj_coeffs_raw) != len(variables):
        raise ValueError(
            f"objective_coefficients has length {len(obj_coeffs_raw)} but "
            f"variables has length {len(variables)}"
        )
    obj_coeffs = [_to_fraction(v, f"objective_coefficients[{i}]")
                  for i, v in enumerate(obj_coeffs_raw)]

    cons_raw = raw["constraints"]
    if not cons_raw:
        raise ValueError("constraints list is empty")
    constraints: list[Constraint] = []
    for i, c in enumerate(cons_raw):
        for k in ("coefficients", "sign", "rhs"):
            if k not in c:
                raise ValueError(f"constraint[{i}]: missing key '{k}'")
        if len(c["coefficients"]) != len(variables):
            raise ValueError(
                f"constraint[{i}]: coefficients length {len(c['coefficients'])} "
                f"!= variables length {len(variables)}"
            )
        sign = c["sign"]
        if sign not in ("<=", "=", ">="):
            raise ValueError(f"constraint[{i}]: sign must be one of <=, =, >=, got '{sign}'")
        coeffs = [_to_fraction(v, f"constraint[{i}].coefficients[{j}]")
                  for j, v in enumerate(c["coefficients"])]
        rhs = _to_fraction(c["rhs"], f"constraint[{i}].rhs")
        # Normalize negative RHS by flipping the row.
        if rhs < 0:
            coeffs = [-x for x in coeffs]
            rhs = -rhs
            sign = {"<=": ">=", ">=": "<=", "=": "="}[sign]
        constraints.append(Constraint(
            name=str(c.get("name", f"R{i+1}")),
            coefficients=coeffs,
            sign=sign,
            rhs=rhs,
        ))

    return Problem(name=name, objective=objective, variables=variables,
                   objective_coefficients=obj_coeffs, constraints=constraints)


# =============================================================================
# Section 4 — Standardisation (Tabla Resumen)
# =============================================================================
@dataclass
class StandardForm:
    """LP put into simplex-ready form."""
    column_labels: list[str]            # ['X1', 'X2', 'S1', 'S2', 'A2', ...]
    cj: list[BigM]                      # objective coefficient per column
    body: list[list[Fraction]]          # constraint matrix (m × n)
    rhs: list[Fraction]                 # length m
    basic: list[int]                    # column index of the initial basic var per row
    artificial_cols: set[int]           # indices of artificial variables (for infeasibility check)


def standardize(problem: Problem) -> StandardForm:
    is_max = problem.objective == "max"
    big_m_sign = -1 if is_max else 1   # max → -M, min → +M

    n_vars = len(problem.variables)
    n_cons = len(problem.constraints)

    # Decision-variable columns first.
    col_labels: list[str] = list(problem.variables)
    cj: list[BigM] = [BigM(c, 0) for c in problem.objective_coefficients]

    # Build the body matrix with zero-padding; we will append slack/surplus/
    # artificial columns as we discover them.
    body: list[list[Fraction]] = [list(c.coefficients) for c in problem.constraints]
    rhs: list[Fraction] = [c.rhs for c in problem.constraints]
    basic: list[int] = [-1] * n_cons
    artificial_cols: set[int] = set()

    def add_column(label: str, cj_value: BigM, entries_per_row: dict[int, Fraction]):
        nonlocal col_labels, cj, body
        col_labels.append(label)
        cj.append(cj_value)
        col_pos = len(col_labels) - 1
        for r in range(n_cons):
            body[r].append(entries_per_row.get(r, Fraction(0)))
        return col_pos

    # Pedagogical convention (matches the reference PDFs): name slack/surplus
    # and artificial variables by the row index where they're introduced — so
    # the artificial in row 2 is A2, the surplus in row 5 is S5, etc.
    for r, con in enumerate(problem.constraints):
        row_label = r + 1
        if con.sign == "<=":
            col = add_column(f"S{row_label}", BigM(0, 0), {r: Fraction(1)})
            basic[r] = col
        elif con.sign == "=":
            col = add_column(f"A{row_label}", BigM(0, big_m_sign), {r: Fraction(1)})
            basic[r] = col
            artificial_cols.add(col)
        else:  # ">="
            add_column(f"S{row_label}", BigM(0, 0), {r: Fraction(-1)})
            col = add_column(f"A{row_label}", BigM(0, big_m_sign), {r: Fraction(1)})
            basic[r] = col
            artificial_cols.add(col)

    return StandardForm(
        column_labels=col_labels,
        cj=cj,
        body=body,
        rhs=rhs,
        basic=basic,
        artificial_cols=artificial_cols,
    )


# =============================================================================
# Section 5 — Tableau snapshots
# =============================================================================
@dataclass
class Snapshot:
    title: str
    column_labels: list[str]
    cj: list[BigM]
    basic: list[int]                                # column index per row
    body: list[list[Fraction]]
    rhs: list[Fraction]
    zj: list[BigM]
    cj_minus_zj: list[BigM]
    z_value: BigM
    ratios: list[Optional[Fraction]] = field(default_factory=list)
    pivot_row: Optional[int] = None
    pivot_col: Optional[int] = None
    row_ops: list[str] = field(default_factory=list)


def _zj_row(sf_or_state, basic, cj, body) -> list[BigM]:
    n_cols = len(cj)
    zj = [BigM(0, 0) for _ in range(n_cols)]
    for r, b_idx in enumerate(basic):
        cb = cj[b_idx]
        for j in range(n_cols):
            zj[j] = zj[j] + cb * body[r][j]
    return zj


def _z_value(basic, cj, rhs) -> BigM:
    z = BigM(0, 0)
    for r, b_idx in enumerate(basic):
        z = z + cj[b_idx] * rhs[r]
    return z


# =============================================================================
# Section 6 — Simplex iterator
# =============================================================================
class SimplexResult:
    OPTIMAL = "optimal"
    UNBOUNDED = "unbounded"
    INFEASIBLE = "infeasible"


@dataclass
class SolveOutcome:
    status: str
    snapshots: list[Snapshot]
    sf: StandardForm
    problem: Problem
    final_basic: list[int]
    final_body: list[list[Fraction]]
    final_rhs: list[Fraction]
    z_value: BigM


def solve(problem: Problem) -> SolveOutcome:
    sf = standardize(problem)
    is_max = problem.objective == "max"

    # Working state (deep-copied between snapshots).
    body = [row[:] for row in sf.body]
    rhs = sf.rhs[:]
    basic = sf.basic[:]
    cj = sf.cj   # cj never changes through pivots
    snapshots: list[Snapshot] = []

    iteration = 0
    # Each row's most recent operation label, indexed by row position.
    row_ops = [""] * len(basic)

    while True:
        zj = _zj_row(sf, basic, cj, body)
        cj_zj = [cj[j] - zj[j] for j in range(len(cj))]
        z_val = _z_value(basic, cj, rhs)

        # Stopping condition.
        if is_max:
            optimal = all(v <= BigM(0) for v in cj_zj)
            best_idx = max(range(len(cj_zj)), key=lambda j: cj_zj[j])
            best_is_improving = cj_zj[best_idx] > BigM(0)
        else:
            optimal = all(v >= BigM(0) for v in cj_zj)
            best_idx = min(range(len(cj_zj)), key=lambda j: cj_zj[j])
            best_is_improving = cj_zj[best_idx] < BigM(0)

        if optimal or not best_is_improving:
            snapshots.append(Snapshot(
                title=f"Iteración {iteration} (óptima)",
                column_labels=sf.column_labels[:],
                cj=cj[:],
                basic=basic[:],
                body=[row[:] for row in body],
                rhs=rhs[:],
                zj=zj,
                cj_minus_zj=cj_zj,
                z_value=z_val,
                ratios=[None] * len(basic),
                pivot_row=None,
                pivot_col=None,
                row_ops=row_ops[:],
            ))
            # Infeasibility check: any artificial still basic with nonzero RHS?
            for r, b_idx in enumerate(basic):
                if b_idx in sf.artificial_cols and rhs[r] != 0:
                    return SolveOutcome(
                        status=SimplexResult.INFEASIBLE,
                        snapshots=snapshots, sf=sf, problem=problem,
                        final_basic=basic, final_body=body, final_rhs=rhs,
                        z_value=z_val,
                    )
            return SolveOutcome(
                status=SimplexResult.OPTIMAL,
                snapshots=snapshots, sf=sf, problem=problem,
                final_basic=basic, final_body=body, final_rhs=rhs,
                z_value=z_val,
            )

        pivot_col = best_idx

        # Min-ratio test (only positive entries in the pivot column).
        ratios: list[Optional[Fraction]] = []
        best_row = -1
        best_ratio: Optional[Fraction] = None
        for r in range(len(basic)):
            entry = body[r][pivot_col]
            if entry > 0:
                ratio = rhs[r] / entry
                ratios.append(ratio)
                if best_ratio is None or ratio < best_ratio or (
                    ratio == best_ratio and basic[r] < basic[best_row]   # Bland's rule on ties
                ):
                    best_ratio = ratio
                    best_row = r
            else:
                ratios.append(None)

        if best_row < 0:
            # Unbounded.
            snapshots.append(Snapshot(
                title=f"Iteración {iteration} (no acotado)",
                column_labels=sf.column_labels[:],
                cj=cj[:],
                basic=basic[:],
                body=[row[:] for row in body],
                rhs=rhs[:],
                zj=zj,
                cj_minus_zj=cj_zj,
                z_value=z_val,
                ratios=ratios,
                pivot_row=None,
                pivot_col=pivot_col,
                row_ops=row_ops[:],
            ))
            return SolveOutcome(
                status=SimplexResult.UNBOUNDED,
                snapshots=snapshots, sf=sf, problem=problem,
                final_basic=basic, final_body=body, final_rhs=rhs,
                z_value=z_val,
            )

        # Snapshot BEFORE the pivot — this matches how the PDFs lay out each table
        # with pivot highlighting and the ratio column.
        snapshots.append(Snapshot(
            title=f"Iteración {iteration}",
            column_labels=sf.column_labels[:],
            cj=cj[:],
            basic=basic[:],
            body=[row[:] for row in body],
            rhs=rhs[:],
            zj=zj,
            cj_minus_zj=cj_zj,
            z_value=z_val,
            ratios=ratios,
            pivot_row=best_row,
            pivot_col=pivot_col,
            row_ops=row_ops[:],
        ))

        # Pivot: divide pivot row by pivot element.
        pivot_val = body[best_row][pivot_col]
        new_pivot_row = [v / pivot_val for v in body[best_row]]
        new_pivot_rhs = rhs[best_row] / pivot_val
        new_row_ops = [""] * len(basic)
        if pivot_val == 1:
            new_row_ops[best_row] = f"F{best_row + 1}"
        else:
            new_row_ops[best_row] = f"({fmt_fraction(Fraction(1) / pivot_val)})·F{best_row + 1}"

        # Eliminate the pivot column from every other row.
        new_body = [row[:] for row in body]
        new_rhs = rhs[:]
        new_body[best_row] = new_pivot_row
        new_rhs[best_row] = new_pivot_rhs

        for r in range(len(basic)):
            if r == best_row:
                continue
            factor = body[r][pivot_col]
            if factor == 0:
                new_row_ops[r] = f"F{r + 1}"
                continue
            new_body[r] = [body[r][j] - factor * new_pivot_row[j] for j in range(len(cj))]
            new_rhs[r] = rhs[r] - factor * new_pivot_rhs
            # Label: e.g. "(-4)·F<pivot> + F<r>"
            sign_factor = -factor
            if sign_factor == 1:
                lhs = f"F{best_row + 1}"
            elif sign_factor == -1:
                lhs = f"-F{best_row + 1}"
            else:
                lhs = f"({fmt_fraction(sign_factor)})·F{best_row + 1}"
            new_row_ops[r] = f"{lhs} + F{r + 1}"

        body = new_body
        rhs = new_rhs
        basic[best_row] = pivot_col
        row_ops = new_row_ops
        iteration += 1


# =============================================================================
# Section 7 — Excel writer
# =============================================================================
THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)

# Neutral grayscale palette with a single soft accent for the pivot element —
# closer to how a student would format an Excel by hand.
FILL_HEADER     = PatternFill("solid", fgColor="F2F2F2")  # encabezado: gris muy claro
FILL_PIVOT_COL  = PatternFill("solid", fgColor="F2F2F2")  # columna pivote: gris muy claro
FILL_PIVOT_ROW  = PatternFill("solid", fgColor="E7E6E6")  # fila pivote: gris claro
FILL_PIVOT_CELL = PatternFill("solid", fgColor="FFE699")  # elemento pivote: amarillo suave
FILL_ZJ         = PatternFill("solid", fgColor="F2F2F2")  # filas Zj / Cj-Zj

FONT_TITLE      = Font(bold=True, size=12)
FONT_SUBTITLE   = Font(bold=True, size=11)
FONT_HEADER     = Font(bold=True)
FONT_BOLD       = Font(bold=True)


NO_BORDER = Border()

# Track each column's longest displayed text so we can size columns correctly
# without measuring the underlying numeric value (a stored 1.333... shouldn't
# blow the column width when the cell will display "1 1/3").
_DISPLAY_WIDTHS: dict[int, int] = {}


def _record_width(col: int, text: str) -> None:
    if text is None:
        return
    cur = _DISPLAY_WIDTHS.get(col, 0)
    if len(text) > cur:
        _DISPLAY_WIDTHS[col] = len(text)


def _set(ws, row, col, value, *, fill=None, font=None, align=CENTER, border=BORDER):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = align
    cell.border = border if border is not None else NO_BORDER
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if isinstance(value, str):
        _record_width(col, value)
    return cell


def _fraction_format(f: Fraction) -> str:
    """Pick an Excel number format that displays this Fraction faithfully.

    - integer       → "0"
    - pure fraction (|num| < denom)  → "??/??;-??/??;0;@"
    - mixed fraction (|num| > denom) → "# ??/??;-# ??/??;0;@"
    The number of '?' placeholders matches the actual denominator's digit
    count, so 17 → '??', 100 → '???'. Excel auto-picks the closest fraction
    with that many denominator digits, so e.g. 882.353… with format
    '# ??/??' renders as '882 6/17'.
    """
    if f.denominator == 1:
        return "0"
    digits = max(len(str(abs(f.denominator))), 1)
    p = "?" * digits
    if abs(f.numerator) < f.denominator:
        return f"{p}/{p};-{p}/{p};0;@"
    return f"# {p}/{p};-# {p}/{p};0;@"


def _set_value(ws, row, col, value, *, fill=None, font=None, align=CENTER, border=BORDER):
    """Write a Fraction / BigM / number / str cell, choosing native numeric
    storage when possible so Excel doesn't flag 'Número almacenado como texto'.

    BigM with an M term has no native numeric form, so it stays as text.
    The displayed string is always recorded for column-width computation.
    """
    if value is None:
        return _set(ws, row, col, "", fill=fill, font=font, align=align, border=border)

    # Unwrap a BigM whose M coefficient is zero — it's just a scalar.
    if isinstance(value, BigM):
        if value.m == 0:
            return _set_value(ws, row, col, value.c,
                              fill=fill, font=font, align=align, border=border)
        text = fmt_bigm(value)
        cell = _set(ws, row, col, text, fill=fill, font=font, align=align, border=border)
        return cell

    if isinstance(value, Fraction):
        display = fmt_fraction(value)
        if value.denominator == 1:
            cell = _set(ws, row, col, int(value.numerator),
                        fill=fill, font=font, align=align, border=border)
        else:
            cell = _set(ws, row, col, float(value),
                        fill=fill, font=font, align=align, border=border)
        cell.number_format = _fraction_format(value)
        _record_width(col, display)
        return cell

    if isinstance(value, int):
        cell = _set(ws, row, col, value, fill=fill, font=font, align=align, border=border)
        cell.number_format = "0"
        _record_width(col, str(value))
        return cell

    if isinstance(value, float):
        cell = _set(ws, row, col, value, fill=fill, font=font, align=align, border=border)
        _record_width(col, str(value))
        return cell

    # Anything else → text path.
    return _set(ws, row, col, str(value), fill=fill, font=font, align=align, border=border)


def write_excel(outcome: SolveOutcome, path: Path) -> None:
    _DISPLAY_WIDTHS.clear()
    wb = Workbook()
    ws = wb.active
    ws.title = "Simplex"

    p = outcome.problem
    sf = outcome.sf
    n_cols = len(sf.column_labels)

    # --- Encabezado: nombre del problema, modelo general y modelo estándar ---
    row = 1
    _set(ws, row, 1, p.name, font=FONT_TITLE, align=LEFT, border=None)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols + 5)
    row += 2

    obj_word = "Maximizar" if p.objective == "max" else "Minimizar"

    # Modelo general — el problema tal como lo planteó el usuario, antes de
    # introducir holguras / artificiales y antes de convertir todo a igualdad.
    _set(ws, row, 1, "Modelo general", font=FONT_SUBTITLE, align=LEFT, border=None)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols + 5)
    row += 1
    gen_obj_terms = _format_linear_expr(
        [(c, lbl) for c, lbl in zip(p.objective_coefficients, p.variables)],
        keep_zero=True,
    )
    _set(ws, row, 1, f"F.O.: {obj_word} Z = {gen_obj_terms}", align=LEFT, border=None)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols + 5)
    row += 1
    _set(ws, row, 1, "Sujeto a:", align=LEFT, border=None)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols + 5)
    row += 1
    for r, con in enumerate(p.constraints):
        terms = _format_linear_expr(
            [(v, lbl) for v, lbl in zip(con.coefficients, p.variables)],
            keep_zero=False,
        )
        _set(ws, row, 1, f"  {r+1}) {terms} {con.sign} {fmt_fraction(con.rhs)}",
             align=LEFT, border=None)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols + 5)
        row += 1
    _set(ws, row, 1, f"  {', '.join(p.variables)} ≥ 0", align=LEFT, border=None)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols + 5)
    row += 2

    _set(ws, row, 1, "Modelo estándar", font=FONT_SUBTITLE, align=LEFT, border=None)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols + 5)
    row += 1

    obj_terms = _format_linear_expr(
        [(c, lbl) for c, lbl in zip(sf.cj, sf.column_labels)],
        keep_zero=True,
    )
    _set(ws, row, 1, f"F.O.: {obj_word} Z = {obj_terms}", align=LEFT, border=None)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols + 5)
    row += 1
    _set(ws, row, 1, "Sujeto a:", align=LEFT, border=None)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols + 5)
    row += 1
    for r, body_row in enumerate(sf.body):
        terms = _format_linear_expr(
            [(v, lbl) for v, lbl in zip(body_row, sf.column_labels)],
            keep_zero=False,
        )
        _set(ws, row, 1, f"  {r+1}) {terms} = {fmt_fraction(sf.rhs[r])}",
             align=LEFT, border=None)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols + 5)
        row += 1
    _set(ws, row, 1, "  Xi, Si, Ai ≥ 0", align=LEFT, border=None)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols + 5)
    row += 2

    # --- Tableau blocks ---
    for snap in outcome.snapshots:
        row = _write_tableau(ws, row, snap) + 1   # blank line between blocks

    # --- Solution / diagnostic block ---
    row = _write_solution_block(ws, row, outcome) + 1

    # --- Column widths (using tracked display lengths, not stored values,
    # so a stored float doesn't inflate the column) ---
    for col_idx, width in _DISPLAY_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(8, min(28, width + 2))

    wb.save(str(path))


def _write_tableau(ws, start_row: int, snap: Snapshot) -> int:
    """Lay out one tableau block starting at start_row. Returns the last row used."""
    n_cols = len(snap.column_labels)

    # Column layout:
    #  col 1: 'Cj' / coeff of basic var
    #  col 2: 'Var. Básicas' / basic var name
    #  cols 3..3+n_cols-1: variable columns
    #  col 3+n_cols: 'b' (RHS)
    #  col 4+n_cols: 'Razón' (ratio)
    #  col 5+n_cols: 'Operación' (row op)
    var_col_start = 3
    var_col_end = var_col_start + n_cols - 1
    b_col = var_col_end + 1
    ratio_col = b_col + 1
    op_col = ratio_col + 1

    r = start_row

    # Título de la tabla (texto plano en negrita, sin relleno).
    _set(ws, r, 1, snap.title, font=FONT_SUBTITLE, align=LEFT, border=None)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=op_col)
    r += 1

    # Header line 1: Cj values (above var columns) + 'b' / 'Razón' / 'Operación'
    _set(ws, r, 1, "Cj", fill=FILL_HEADER, font=FONT_HEADER)
    _set(ws, r, 2, "", fill=FILL_HEADER)
    for j in range(n_cols):
        _set_value(ws, r, var_col_start + j, snap.cj[j],
                   fill=FILL_HEADER, font=FONT_HEADER)
    _set(ws, r, b_col, "", fill=FILL_HEADER)
    _set(ws, r, ratio_col, "", fill=FILL_HEADER)
    _set(ws, r, op_col, "", fill=FILL_HEADER)
    r += 1

    # Header line 2: 'Var. Básicas' + variable names + 'b' / 'Razón' / 'Operación'
    _set(ws, r, 1, "Cj", fill=FILL_HEADER, font=FONT_HEADER)
    _set(ws, r, 2, "Var. Básicas", fill=FILL_HEADER, font=FONT_HEADER)
    for j, lbl in enumerate(snap.column_labels):
        _set(ws, r, var_col_start + j, lbl, fill=FILL_HEADER, font=FONT_HEADER)
    _set(ws, r, b_col, "b", fill=FILL_HEADER, font=FONT_HEADER)
    _set(ws, r, ratio_col, "Razón", fill=FILL_HEADER, font=FONT_HEADER)
    _set(ws, r, op_col, "Operación", fill=FILL_HEADER, font=FONT_HEADER)
    r += 1

    # Data rows (one per constraint)
    for row_idx, b_idx in enumerate(snap.basic):
        is_pivot_row = (snap.pivot_row == row_idx)
        # Cj of basic var
        _set_value(ws, r, 1, snap.cj[b_idx],
                   fill=FILL_PIVOT_ROW if is_pivot_row else None)
        # Basic var name
        _set(ws, r, 2, snap.column_labels[b_idx],
             fill=FILL_PIVOT_ROW if is_pivot_row else None,
             font=FONT_HEADER)
        # Body coefficients
        for j in range(n_cols):
            is_pivot_col = (snap.pivot_col == j)
            if is_pivot_row and is_pivot_col:
                fill = FILL_PIVOT_CELL
            elif is_pivot_row:
                fill = FILL_PIVOT_ROW
            elif is_pivot_col:
                fill = FILL_PIVOT_COL
            else:
                fill = None
            _set_value(ws, r, var_col_start + j,
                       snap.body[row_idx][j], fill=fill)
        # RHS
        _set_value(ws, r, b_col, snap.rhs[row_idx],
                   fill=FILL_PIVOT_ROW if is_pivot_row else None)
        # Ratio (None → empty cell)
        ratio = snap.ratios[row_idx] if snap.ratios else None
        if ratio is None:
            _set(ws, r, ratio_col, "",
                 fill=FILL_PIVOT_ROW if is_pivot_row else None)
        else:
            _set_value(ws, r, ratio_col, ratio,
                       fill=FILL_PIVOT_ROW if is_pivot_row else None)
        # Operation label (always text)
        _set(ws, r, op_col, snap.row_ops[row_idx] if snap.row_ops else "",
             align=Alignment(horizontal="left", vertical="center"))
        r += 1

    # Zj row
    _set(ws, r, 1, "", fill=FILL_ZJ)
    _set(ws, r, 2, "Zj", fill=FILL_ZJ, font=FONT_HEADER)
    for j in range(n_cols):
        is_pivot_col = (snap.pivot_col == j)
        _set_value(ws, r, var_col_start + j, snap.zj[j],
                   fill=FILL_PIVOT_COL if is_pivot_col else FILL_ZJ)
    _set_value(ws, r, b_col, snap.z_value, fill=FILL_ZJ, font=FONT_HEADER)
    _set(ws, r, ratio_col, "", fill=FILL_ZJ)
    _set(ws, r, op_col, "", fill=FILL_ZJ)
    r += 1

    # Cj - Zj row
    _set(ws, r, 1, "", fill=FILL_ZJ)
    _set(ws, r, 2, "Cj - Zj", fill=FILL_ZJ, font=FONT_HEADER)
    for j in range(n_cols):
        is_pivot_col = (snap.pivot_col == j)
        _set_value(ws, r, var_col_start + j, snap.cj_minus_zj[j],
                   fill=FILL_PIVOT_COL if is_pivot_col else FILL_ZJ,
                   font=FONT_BOLD if is_pivot_col else None)
    _set(ws, r, b_col, "", fill=FILL_ZJ)
    _set(ws, r, ratio_col, "", fill=FILL_ZJ)
    _set(ws, r, op_col, "", fill=FILL_ZJ)
    return r


def _write_solution_block(ws, start_row: int, outcome: SolveOutcome) -> int:
    sf = outcome.sf
    n_cols = len(sf.column_labels)
    last_col = n_cols + 5

    r = start_row
    if outcome.status == SimplexResult.OPTIMAL:
        title = "Solución óptima"
    elif outcome.status == SimplexResult.UNBOUNDED:
        title = "Solución no acotada"
    else:
        title = "Problema infactible"

    _set(ws, r, 1, title, font=FONT_SUBTITLE, align=LEFT, border=None)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
    r += 1

    if outcome.status == SimplexResult.OPTIMAL:
        # Cada variable básica toma su valor de la columna b; el resto vale 0.
        values: dict[str, Fraction] = {lbl: Fraction(0) for lbl in sf.column_labels}
        for row_idx, b_idx in enumerate(outcome.final_basic):
            values[sf.column_labels[b_idx]] = outcome.final_rhs[row_idx]

        for lbl in sf.column_labels:
            _set(ws, r, 1, lbl, font=FONT_BOLD, align=LEFT, border=None)
            _set(ws, r, 2, "=", align=LEFT, border=None)
            _set_value(ws, r, 3, values[lbl], align=LEFT, border=None)
            r += 1

        _set(ws, r, 1, "Z", font=FONT_BOLD, align=LEFT, border=None)
        _set(ws, r, 2, "=", align=LEFT, border=None)
        _set_value(ws, r, 3, outcome.z_value, font=FONT_BOLD, align=LEFT, border=None)
        r += 1
    elif outcome.status == SimplexResult.UNBOUNDED:
        msg = ("La columna entrante no tiene entradas positivas, "
               "así que la función objetivo no está acotada.")
        _set(ws, r, 1, msg, align=LEFT, border=None)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        r += 1
    else:  # INFEASIBLE
        msg = ("Quedó al menos una variable artificial en la base con valor "
               "positivo, así que el problema no tiene solución factible.")
        _set(ws, r, 1, msg, align=LEFT, border=None)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=last_col)
        r += 1
    return r


# =============================================================================
# Section 8 — CLI entry point
# =============================================================================
def main(argv: list[str]) -> int:
    in_path = Path(argv[1]) if len(argv) >= 2 else Path("problem.json")
    out_path = Path(argv[2]) if len(argv) >= 3 else Path("output.xlsx")

    if not in_path.exists():
        print(f"error: input file not found: {in_path}", file=sys.stderr)
        return 2

    try:
        problem = load_problem(in_path)
    except (ValueError, json.JSONDecodeError) as e:
        print(f"error parsing {in_path}: {e}", file=sys.stderr)
        return 2

    outcome = solve(problem)
    write_excel(outcome, out_path)

    # One-line stdout summary.
    if outcome.status == SimplexResult.OPTIMAL:
        basic_summary = ", ".join(
            f"{outcome.sf.column_labels[b]}={fmt_fraction(outcome.final_rhs[r])}"
            for r, b in enumerate(outcome.final_basic)
        )
        print(f"OK — Z = {fmt_bigm(outcome.z_value)}  ({basic_summary})  →  {out_path}")
    elif outcome.status == SimplexResult.UNBOUNDED:
        print(f"UNBOUNDED — see {out_path}")
    else:
        print(f"INFEASIBLE — see {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
